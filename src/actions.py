from PySide.QtCore import *
from PySide.QtGui import *
from models import *
#import xlwt
from openpyxl import Workbook

class Actions(object):
    
    def __init__(self,mainWindow):

        self.mainWindow=mainWindow
        
        self.basicActions=QActionGroup(self.mainWindow)        
        self.actionOpen=self.basicActions.addAction(QIcon(":/icons/data/icons/document-import.png"),"Open Database")        
        self.actionNew=self.basicActions.addAction(QIcon(":/icons/data/icons/window_new.png"),"New Database")        
                
        self.databaseActions=QActionGroup(self.mainWindow)
        self.actionExport=self.databaseActions.addAction(QIcon(":icons/data/icons/document-export.png"),"Export Data")        
        self.actionAdd=self.databaseActions.addAction(QIcon(":/icons/data/icons/bookmark_add.png"),"Add Nodes")                
        self.actionDelete=self.databaseActions.addAction(QIcon(":/icons/data/icons/editdelete.png"),"Delete Selected Nodes")
        
        self.facebookActions=QActionGroup(self.mainWindow)        
        self.actionQuery=self.facebookActions.addAction(QIcon(":/icons/data/icons/find.png"),"Query")                
        self.actionShowColumns=self.facebookActions.addAction("Show Columns")
        self.actionUnpackData=self.facebookActions.addAction("Unpack Data")

        #connect the actions to their corresponding action functions (slots)
        self.actionOpen.triggered.connect(self.openDB)
        self.actionNew.triggered.connect(self.makeDB)
        self.actionExport.triggered.connect(self.exportNodes)
        self.actionAdd.triggered.connect(self.addNodes)        
        self.actionDelete.triggered.connect(self.deleteNodes)
        self.actionQuery.triggered.connect(self.queryNodes)
        self.actionShowColumns.triggered.connect(self.showColumns)
        self.actionUnpackData.triggered.connect(self.unpackData)
        
    @Slot()
    def openDB(self):
        #open a file dialog with a .db filter
        fldg=QFileDialog(caption="Open DB File",directory=self.mainWindow.settings.value("lastpath","."),filter="DB files (*.db)")
        fldg.setFileMode(QFileDialog.ExistingFile)
        if fldg.exec_():
            self.mainWindow.database.connect(fldg.selectedFiles()[0])           
            self.mainWindow.settings.setValue("lastpath",fldg.selectedFiles()[0])                            
            self.mainWindow.updateUI()
            self.mainWindow.treemodel.reset()  

            
    @Slot()
    def makeDB(self):
        #same as openDB-Slot, but now for creating a new one on the file system
        fldg=QFileDialog(caption="Save DB File",directory=self.mainWindow.settings.value("lastpath","."),filter="DB files (*.db)")        
        fldg.setAcceptMode(QFileDialog.AcceptSave)
        fldg.setDefaultSuffix("db")
               
        if fldg.exec_():
            self.mainWindow.database.createconnect(fldg.selectedFiles()[0])
            self.mainWindow.settings.setValue("lastpath",fldg.selectedFiles()[0])
            self.mainWindow.updateUI()
            self.mainWindow.treemodel.reset()

        
    @Slot()
    def deleteNodes(self):   
        progress = QProgressDialog("Deleting data...", "Abort", 0, 0,self.mainWindow)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.forceShow()     
                             
        todo=self.mainWindow.tree.selectedIndexesAndChildren(None,True)
        progress.setMaximum(len(todo))            
        
        c=0
        for index in todo:            
            progress.setValue(c)
            c+=1            
            
            self.mainWindow.treemodel.deleteNode(index)           
           
            if progress.wasCanceled():
                break 
        progress.cancel()

                     
                                 
    @Slot()
    def exportNodes(self):
                                        
        fldg=QFileDialog(caption="Export DB File to XLSX",filter="XLS Files (*.xlsx)")
        fldg.setAcceptMode(QFileDialog.AcceptSave)
        fldg.setDefaultSuffix("xlsx")
        
        if fldg.exec_():               
            progress = QProgressDialog("Saving data...", "Abort", 0, 0,self.mainWindow)
            progress.setWindowModality(Qt.WindowModal)
            progress.setMinimumDuration(0)
            progress.forceShow()   
             
            
                   
            #selected=self.mainWindow.tree.selectedIndexesAndChildren()
            progress.setMaximum(Node.query.count())
            
                    
            if os.path.isfile(fldg.selectedFiles()[0]):
                os.remove(fldg.selectedFiles()[0])
          
            
            #x=xlwt.Workbook(encoding="latin-1")
            x=Workbook(encoding="latin-1",optimized_write = True)
            
            
            def writeValue(sheet,row,col,value):
                sheet.cell(row=row,column=col).value=value
                
            sheets={}
            def getSheet(level=1):
                sh = sheets.get(level,None)
                if sh==None:                  
                    #xs=x.add_sheet("Level"+str(level))  
                    xs=x.create_sheet()
                    xs.title="Level"+str(level)
                    
                    row=["level","id","parent_id","facebook_id","query_status","query_time","query_type","response"]
                    for key in self.mainWindow.treemodel.customcolumns:
                        row.append(key)
                        
                    xs.append(row)
                                                         
#                    writeValue(xs,0,0,"level")
#                    writeValue(xs,0,1,"id")
#                    writeValue(xs,0,2,"parent_id")                
#                    writeValue(xs,0,3,"facebook id")
#                    writeValue(xs,0,4,"query status")
#                    writeValue(xs,0,5,"query time")
#                    writeValue(xs,0,6,"query type")
#                    writeValue(xs,0,7,"response")
#                    col=8    
#                    for key in self.mainWindow.treemodel.customcolumns:
#                        writeValue(xs,0,col,key)
#                        col+=1                     
                    
                    sh={'sheet':xs,'row':1}
                    sheets[level]=sh
                return sh                    
            
            page=0
            no=0  
            while True:            
                allnodes=Node.query.offset(page*5000).limit(5000).all()

                for node in allnodes:      
                    if progress.wasCanceled():
                        break
                               
                    progress.setValue(no)
                    no+=1            
                   

                                       
                    sh=getSheet(node.level)
                    xs=sh['sheet']
                    #row=sh['row']
                    #sh['row']+=1
                    
                    row=[node.level,node.id,node.parent_id,node.facebookid,node.querystatus,node.querytime,node.querytype,node.response_raw]    
                    for key in self.mainWindow.treemodel.customcolumns:                    
                        row.append(node.getResponseValue(key))    
                     
                    xs.append(row)                                 
#                    writeValue(xs,row,0,node.level)
#                    writeValue(xs,row,1,node.id)
#                    writeValue(xs,row,2,node.parent_id)
#                    writeValue(xs,row,3,node.facebookid)
#                    writeValue(xs,row,4,node.querystatus)
#                    writeValue(xs,row,5,node.querytime)
#                    writeValue(xs,row,6,node.querytype)
#                    writeValue(xs,row,7,node.response_raw)
                            
#                    col=8    
#                    for key in self.mainWindow.treemodel.customcolumns:                    
#                        value=node.getResponseValue(key)
#                        writeValue(xs,row,col,value)
    #                    try:
    #                        if len(unicode(key))<32760:
    #                            xs.write(row,col,unicode(value))
    #                        else:
    #                            xs.write(row,col,"ERROR")
    #                    except:
    #                        xs.write(row,col,"ERROR")
    
                        #col+=1
                 
                if progress.wasCanceled():
                    break
                           
                if len(allnodes) == 0:
                    break
                else:
                    page +=1                        
                                      
                              
            x.remove_sheet(x.get_active_sheet())                  
            x.save(fldg.selectedFiles()[0])
            progress.cancel()

    

    @Slot()
    def addNodes(self):
        if not self.mainWindow.database.connected:
            return False
        
        # makes the user add a new facebook object into the db
        dialog=QDialog(self.mainWindow)
        dialog.setWindowTitle("Add Facebook Objects")
        layout=QVBoxLayout()
        
        label=QLabel("<b>Facebook IDs (one ID per line):</b>")
        layout.addWidget(label)
        
        input=QTextEdit()           
        input.setMinimumWidth(500)
        input.LineWrapMode=QTextEdit.NoWrap
        input.acceptRichText=False
        input.setFocus()
        layout.addWidget(input)
                
        
        buttons=QDialogButtonBox(QDialogButtonBox.Ok|QDialogButtonBox.Cancel)
        layout.addWidget(buttons)
        
        dialog.setLayout(layout)
        
        def createNodes():
            self.mainWindow.treemodel.addNodes(input.toPlainText().splitlines())
            dialog.close()  

        def close():
            dialog.close()
            
        #connect the nested functions above to the dialog-buttons
        buttons.accepted.connect(createNodes)
        buttons.rejected.connect(close)
        dialog.exec_()
        
        
    @Slot()     
    def showColumns(self):
        self.mainWindow.treemodel.setCustomColumns(self.mainWindow.fieldList.toPlainText().splitlines())
                                        
    @Slot()     
    def unpackData(self):
        pass
  
    @Slot()
    def queryNodes(self):
        #Show progress window         
        progress = QProgressDialog("Fetching data...", "Abort", 0, 0,self.mainWindow)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.forceShow()       
                                
        #Get selected nodes
        level=self.mainWindow.levelEdit.value()
        todo=self.mainWindow.tree.selectedIndexesAndChildren(level)
        progress.setMaximum(len(todo))
            

        #Set options
        relation=self.mainWindow.relationEdit.currentText()
        appendempty=self.mainWindow.emptyEdit.isChecked()
        options={}
        options['since']=self.mainWindow.sinceEdit.date().toString("yyyy-MM-dd")
        options['until']=self.mainWindow.untilEdit.date().toString("yyyy-MM-dd")
        options['offset']=self.mainWindow.offsetEdit.value()
        options['limit']=self.mainWindow.limitEdit.value()
        
        #delete old data
        if self.mainWindow.deleteEdit.isChecked():
            self.mainWindow.treemodel.delete(level,relation)
        
        #Fetch data
        c=0
        for index in todo:            
            progress.setValue(c)
            c+=1            
            
            self.mainWindow.treemodel.queryData(index,relation,options,appendempty)           
           
            if progress.wasCanceled():
                break 
 

        progress.cancel()   

        
 
           